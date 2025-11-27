from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


def process_datasets(number, file1, file2):
    wb1 = load_workbook(file1)
    ws1 = wb1['Sheet1']  # 明确指定工作表名称
    wb2 = load_workbook(file2)
    ws2 = wb2['Sheet1']

    # 初始化统计变量
    counter_stats = {5: 0, 6: 0, 7: 0}
    column_stats = {3: 0, 5: 0, 7: 0, 9: 0, 11: 0, 13: 0, 15: 0}  # C/E/G/I/K/M/O列
    comparison_pairs = [(3, 4), (5, 6), (7, 8), (9, 10), (11, 12), (13, 14), (15, 16)]  # 完整列对列表

    # 遍历所有数据集（假设有n个数据集）
    for data_set_index in range(1, 146):  # 示例处理16个数据集
        # 计算起始坐标
        base_row = 3 + 5 * (data_set_index - 1)
        base_col = 3

        # 新的D列更新行
        base_row_d = 2 + 5 * (data_set_index - 1)

        # 初始化计数器
        counter = 0

        skip_dataset = False  # 新增跳过标志

        # 处理C(3)/E(5)/G(7)/I(9)/K(11)/M(13)/O(15)列
        for col_offset in [0, 2, 4, 6, 8, 10, 12]:
            current_col = base_col + col_offset

            # 读取file2的后四行指标值（行号base_row+3到base_row+6）（新增空值检查）
            try:
                file2_values = [
                    float(ws2.cell(row=base_row + i, column=current_col).value)
                    for i in range(4)
                ]
            except (TypeError, ValueError):
                # 遇到空值或非数值时跳过该数据集
                skip_dataset = True
                break  # 跳出列遍历循环

            if skip_dataset:
                continue  # 跳过当前数据集的后续处理

            # 读取file1对应的基准值（前三个指标）
            file1_ref = [
                ws1.cell(row=base_row + i, column=current_col).value
                for i in range(3)
            ]

            # 指标比较（前三个值中需≥2个达标）
            pass_condition = sum(
                f2 >= f1 for f2, f1 in zip(file2_values[:3], file1_ref)
            ) >= 2

            # start
            # 判断前一列是否性能优于后一列
            col1, col2 = current_col, current_col+1
            # 获取两列的前三个指标值
            col1_values = [
                ws2.cell(row=base_row + i, column=col1).value
                for i in range(3)
            ]
            col2_values = [
                ws2.cell(row=base_row + i, column=col2).value
                for i in range(3)
            ]

            # 逐个指标比较（前三个指标）
            comparison_results = []
            for v1, v2 in zip(col1_values, col2_values):
                try:
                    # 转换为浮点数比较（根据上下文数据模式[0][1]）
                    comparison_results.append(float(v1) >= float(v2))
                except (TypeError, ValueError):
                    # 处理空值/非数值情况
                    comparison_results.append(False)
            # end

            # 判断条件：不归一的优于归一的且不归一的本身性能优越
            if pass_condition and sum(comparison_results) >= 2 and ws2.cell(row=base_row + 1, column=col1).value >= 0.2:
                # 更新file1当前列和下一列
                for i in range(4):
                    # 更新当前列
                    ws1.cell(row=base_row + i, column=current_col,
                             value=file2_values[i])
                    # 更新相邻列
                    adj_value = ws2.cell(
                        row=base_row + i,
                        column=current_col + 1
                    ).value
                    ws1.cell(row=base_row + i,
                             column=current_col + 1,
                             value=adj_value)

        if skip_dataset:
            print(f"数据集 {data_set_index} 包含空值，已跳过")
            continue  # 跳过当前数据集的后续处理

        # 列对比较逻辑（每处理完两个相邻列）
        for pair in comparison_pairs:
            col1, col2 = pair
            # 获取两列的前三个指标值
            col1_values = [
                ws1.cell(row=base_row + i, column=col1).value
                for i in range(3)
            ]
            col2_values = [
                ws1.cell(row=base_row + i, column=col2).value
                for i in range(3)
            ]

            # 逐个指标比较（前三个指标）
            comparison_results = []
            for v1, v2 in zip(col1_values, col2_values):
                try:
                    # 转换为浮点数比较（根据上下文数据模式[0][1]）
                    comparison_results.append(float(v1) >= float(v2))
                except (TypeError, ValueError):
                    # 处理空值/非数值情况
                    comparison_results.append(False)

            # 判断条件：至少两个指标达标
            if sum(comparison_results) >= 2 and ws1.cell(row=base_row + 1, column=col1).value >= 0.2:
                counter += 1
                # 记录列更新次数
                column_stats[col1] += 1

                # 可选：标记特殊格式（根据上下文中的高值模式[1]）
                for i in range(3):
                    ws1.cell(base_row + i, col1).fill = PatternFill(fgColor="FFC7CE", fill_type="solid")

            # 更新D列值（替换原A列更新）
            # d_cell_value = ws1.cell(row=base_row_d, column=4).value or ""
            # clean_value = str(d_cell_value).split('(')[0].strip()
            # ws1.cell(row=base_row_d, column=4, value=f"{clean_value}({counter})")
            # 更新D列值（红色标记条件）
            d_cell = ws1.cell(row=base_row_d, column=4)
            d_cell.value = counter
            if counter >= 5:
                d_cell.font = Font(color="FF0000")  # 红色字体

        # 统计counter次数
        if 5 <= counter <= 7:
            counter_stats[counter] += 1

        print(f"数据集 {data_set_index} 已更新")

    # 输出统计结果
    print("Counter统计结果：")
    print(f"等于5的次数：{counter_stats[5]}")
    print(f"等于6的次数：{counter_stats[6]}")
    print(f"等于7的次数：{counter_stats[7]}")

    print("\n各列触发更新次数：")
    for col in [3, 5, 7, 9, 11, 13, 15]:
        print(f"列{chr(64 + col)}: {column_stats[col]}次")

    wb1.save(f'D:\指定轮数聚类结果(更新后)/{number}轮updated.xlsx')

num = 0
process_datasets(
    number=num,
    file1=f'D:\指定轮数聚类结果(归一化)/{num}轮聚类结果(归一化).xlsx',
    file2=f'D:/指定轮数聚类结果(未归一化)/{num}轮聚类结果(未归一化).xlsx'
)