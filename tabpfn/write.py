import pandas as pd
from openpyxl import load_workbook


def write_transposed_results_to_excel(df_results,  data_set_index):
    """
    将聚类结果表格的特定数据集（比如第16个）转置后写入Excel文件。

    Parameters:
    - df_results: 包含聚类结果的DataFrame
    - file_path: 要写入的Excel文件路径
    - sheet_name: 要写入的工作表名称
    - data_set_index: 数据集索引（例如第16个数据集，传入15）
    - start_row: 写入数据的起始行
    - start_col: 写入数据的起始列
    """
    file_path = 'D:/归一化17轮.xlsx'
    sheet_name = 'Sheet1'
    start_row = 3 + 5 * (data_set_index - 1)
    start_col = 3
    # 选择特定数据集（0基索引，所以是15对应第16个数据集）
    selected_data = df_results.iloc[:, 1:]  # 选择Accuracy, NMI, ARI的行

    # 转置数据
    transposed_data = selected_data.transpose()

    # 打开Excel文件，准备写入
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        workbook = writer.book
        sheet = workbook[sheet_name]

        # 将转置后的数据写入Excel文件
        for i in range(transposed_data.shape[0]):  # 行
            for j in range(transposed_data.shape[1]):  # 列
                value = transposed_data.iloc[i, j]
                # 将值限制为小数点后四位
                rounded_value = round(value, 4)
                sheet.cell(row=start_row + i, column=start_col + 2 * j, value=rounded_value)

    # 保存修改后的Excel文件
    print(f"Data written successfully to {file_path}")

def write_test_tree_to_excel(numerator, denominator, data_set_index):
    file_path = 'D:/tabpfn实验数据.xlsx'
    sheet_name = 'Sheet1'
    start_row = 2 + 4 * (data_set_index - 1)
    start_col = 3

    # 打开Excel文件，准备写入
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        workbook = writer.book
        sheet = workbook[sheet_name]

        numerator = round(numerator, 4)
        denominator = round(denominator, 4)
        # 将格式化为 "分子/分母" 的字符串写入单元格
        cell_value = f"{numerator}/{denominator}"
        sheet.cell(row=start_row, column=start_col, value=cell_value)

    # 保存修改后的Excel文件
    print(f"Data {data_set_index} written successfully to {file_path}")